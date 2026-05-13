#!/usr/bin/env python3
"""
Utilidades para importación de datos desde archivos Excel.
"""

import logging
import re
from datetime import date, datetime, time

import pandas as pd
from openpyxl import load_workbook

from crue_traslados.models import Traslado

logger = logging.getLogger (__name__)


# -------------------------------------------------
# Parse hour helper
# -------------------------------------------------

def parse_hora (hora):
	"""Parsea un valor de hora (time, string) y retorna tupla (hora, minuto)."""
	try:
		# Already datetime.time
		if isinstance (hora, time):
			return hora.hour, hora.minute

		txt = str (hora).strip ()
		txt = txt.replace (".", ":")

		formatos = [
			"%H:%M",
			"%H:%M:%S",
		]

		for fmt in formatos:
			try:
				t = datetime.strptime (txt, fmt)
				return t.hour, t.minute
			except Exception:
				pass
		raise ValueError (f"Formato hora erronea: {hora}")
	except Exception as ex:
		logger.warning ("Error formateando hora: %s — %s", hora, ex)
		return ""


# ---------------------------------------------------------------------------------
# ---------------------------------------------------------------------------------

def parse_fechas_horas (fecha, horaReporte, horaEgreso, horaIngreso):
	"""
	Parsea fecha y horas desde valores de Excel.

	fecha:
		Excel datetime/string like: '2026-04-01 00:00:00'
	horas:
		Excel time/string like: '07:55:00' '7:55' '7.55'

	Returns: (dt_reporte, dt_egreso, dt_ingreso)
	"""
	try:
		fecha_date = None
		# Check what 'horas' are available
		if not horaReporte and horaEgreso:
			horaReporte = horaEgreso

		# datetime object
		if isinstance (fecha, datetime):
			fecha_date = fecha.date ()

		# date object
		elif isinstance (fecha, date):
			fecha_date = fecha
		else:
			txt = str (fecha).strip ()
			# Try ISO excel format first
			try:
				fecha_date = datetime.strptime (
					txt,
					"%Y-%m-%d %H:%M:%S"
				).date ()
			except Exception:
				pass

			# Fallbacks
			if fecha_date is None:
				candidatos = re.findall (
					r"\d{1,2}/+\d{1,2}/+\d{4}",
					txt
				)
				cand = re.sub (r"/+", "/", candidatos [0])
				partes = cand.split ("/")
				a, b, y = partes
				dt = None
				# MM/DD/YYYY
				try:
					dt = datetime.strptime (
						f"{a}/{b}/{y}",
						"%m/%d/%Y"
					)
				except Exception:
					pass

				# DD/MM/YYYY
				if dt is None:
					try:
						dt = datetime.strptime (
							f"{a}/{b}/{y}",
							"%d/%m/%Y"
						)
					except Exception:
						pass

				fecha_date = dt.date ()
	except Exception as ex:
		logger.warning ("Error formateando fecha desde %s — %s", fecha, ex)
		fecha_date = None

	# -------------------------------------------------
	# Parse times
	# -------------------------------------------------

	hr_rep = parse_hora (horaReporte)
	hr_egr = parse_hora (horaEgreso)
	hr_ing = parse_hora (horaIngreso)

	# -------------------------------------------------
	# Build datetimes
	# -------------------------------------------------

	try:
		dt_reporte = datetime.combine (fecha_date, datetime.min.time ()).replace (
			hour=hr_rep [0], minute=hr_rep [1]
		)
	except Exception:
		dt_reporte = None

	try:
		dt_egreso = datetime.combine (fecha_date, datetime.min.time ()).replace (
			hour=hr_egr [0], minute=hr_egr [1]
		)
	except Exception:
		dt_egreso = None

	try:
		dt_ingreso = datetime.combine (fecha_date, datetime.min.time ()).replace (
			hour=hr_ing [0], minute=hr_ing [1]
		)
	except Exception:
		dt_ingreso = None

	return (dt_reporte, dt_egreso, dt_ingreso)


# --------------------------------------------------------------------
# Obtener hojas de un archivo Excel
# --------------------------------------------------------------------

def obtenerHojasExcel (file_obj):
	"""
	Retorna la lista de nombres de hojas de un archivo Excel.

	Args:
		file_obj: Ruta al archivo (str) o file-like object (BytesIO).

	Returns:
		Lista de strings con los nombres de las hojas.
	"""
	wb = load_workbook (file_obj, read_only=True, data_only=True)
	hojas = wb.sheetnames
	wb.close ()
	return hojas


# --------------------------------------------------------------------
# Importar traslados desde Excel
# --------------------------------------------------------------------

def import_traslados_from_excel (file_path=None, file_obj=None, sheet_name="ABRIL"):
	"""
	Importa registros de traslados desde un archivo Excel.

	Soporta recibir una ruta de archivo (file_path) o un file-like object (file_obj).
	Antes de insertar, verifica duplicados por fecha_reporte + documento.

	Args:
		file_path: Ruta al archivo Excel (str, opcional).
		file_obj: Objeto file-like (BytesIO, opcional).
		sheet_name: Nombre de la hoja a leer.

	Returns:
		Tupla (insertados, omitidos) con la cantidad de registros procesados.
	"""
	skipN = 7  # Number of lines to skip containing excel header

	fuente = file_path if file_path else file_obj
	if fuente is None:
		raise ValueError ("Debe proporcionar file_path o file_obj.")

	wb = load_workbook (fuente, data_only=True)
	ws = wb [sheet_name]

	registros = []
	omitidos = 0

	for row in ws.iter_rows (min_row=skipN + 2):  # +2 because Excel starts at 1
		try:
			values = []
			for cell in row:
				values.append (str (cell.value) if cell.value is not None else '')

			fechaReporte, fechaEgreso, fechaIngreso = parse_fechas_horas (
				values [0], values [1], values [2], values [3]
			)

			# Verificar duplicado por fecha_reporte + documento
			documento = str (values [5]).strip ()
			if fechaReporte and documento:
				existe = Traslado.objects.filter (
					fecha_reporte=fechaReporte,
					documento=documento,
				).exists ()
				if existe:
					omitidos += 1
					continue

			obj = Traslado (
				fecha_reporte=fechaReporte,
				fecha_egreso=fechaEgreso,
				fecha_ingreso=fechaIngreso,
				nombre_paciente=str (values [4]).strip (),
				documento=documento,
				servicio=str (values [6]).strip (),
				quien_reporta=str (values [7]).strip (),
				destino=str (values [8]).strip (),
				procedimiento=str (values [9]).strip (),
				medico=str (values [10]).strip (),
				aux_enfermeria=str (values [11]).strip (),
				conductor=str (values [12]).strip (),
				radio_operador=str (values [13]).strip (),
				ambulancia=str (values [14]).strip (),
				observacion="" if pd.isna (values [15]) else str (values [15]).strip (),
				mes=fechaReporte.month,
			)
		except Exception as ex:
			logger.error ("Error leyendo archivo excel, fila: %s — %s", row, ex)
			continue

		registros.append (obj)

	wb.close ()

	# Bulk insert
	Traslado.objects.bulk_create (registros, batch_size=500)
	insertados = len (registros)

	logger.info (
		"Importación completada: %d insertados, %d omitidos (hoja: %s)",
		insertados, omitidos, sheet_name
	)

	return (insertados, omitidos)
