#!/usr/bin/env python3

import re
import pandas as pd
from datetime import datetime, time
import pandas as pd
import traceback


# -------------------------------------------------
# Parse hour helper
# -------------------------------------------------
def parse_hora(hora):
	try:
		# Already datetime.time
		if isinstance(hora, time):
			return hora.hour, hora.minute

		txt = str(hora).strip()
		txt = txt.replace(".", ":")

		formatos = [
			"%H:%M",
			"%H:%M:%S",
		]

		for fmt in formatos:
			try:
				t = datetime.strptime(txt, fmt)
				return t.hour, t.minute
			except:
				pass
		raise Exception(f"Formato hora erronea : {hora}")
	except Exception as ex:
		 print (f"+++ Error formateando hora : {hora=}, {ex=}")
		 return ""

#---------------------------------------------------------------------------------
#---------------------------------------------------------------------------------

import re
from datetime import datetime, date, time

def parse_fechas_horas(fecha, horaReporte, horaEgreso, horaIngreso):
	"""
	fecha:
		Excel datetime/string like: '2026-04-01 00:00:00'
	horas:
		Excel time/string like: '07:55:00' '7:55' '7.55'

	Returns: ( dt_reporte, dt_egreso, dt_ingreso) 
	"""
	try: 
		fecha_date = None
		# Check what 'horas' are available
		if not horaReporte and horaEgreso:
			horaReporte = horaEgreso

		# datetime object
		if isinstance(fecha, datetime):
			fecha_date = fecha.date()

		# date object
		elif isinstance(fecha, date):
			fecha_date = fecha
		else:
			txt = str(fecha).strip()
			# Try ISO excel format first
			try:
				fecha_date = datetime.strptime(
					txt,
					"%Y-%m-%d %H:%M:%S"
				).date()
			except:
				pass

			# Fallbacks
			if fecha_date is None:
				candidatos = re.findall(
					r"\d{1,2}/+\d{1,2}/+\d{4}",
					txt
				)
				cand = re.sub(r"/+", "/", candidatos[0])
				partes = cand.split("/")
				a, b, y = partes
				dt = None
				# MM/DD/YYYY
				try:
					dt = datetime.strptime(
						f"{a}/{b}/{y}",
						"%m/%d/%Y"
					)
				except:
					pass

				# DD/MM/YYYY
				if dt is None:
					try:
						dt = datetime.strptime(
							f"{a}/{b}/{y}",
							"%d/%m/%Y"
						)
					except:
						pass

				fecha_date = dt.date()
	except Exception as ex:
		print (f"+++ Error formateando fecha desde {fecha=}, {ex=}")
		fecha_date = ""

	# -------------------------------------------------
	# Parse times
	# -------------------------------------------------

	hr_rep = parse_hora(horaReporte)
	hr_egr = parse_hora(horaEgreso)
	hr_ing = parse_hora(horaIngreso)

	# -------------------------------------------------
	# Build datetimes
	# -------------------------------------------------

	try:
		dt_reporte = datetime.combine( fecha_date, datetime.min.time()).replace(
			hour=hr_rep[0], minute=hr_rep[1]
		)
	except:
		dt_reporte = None

	try:
		dt_egreso = datetime.combine( fecha_date, datetime.min.time()).replace(
			hour=hr_egr[0], minute=hr_egr[1]
		)
	except:
		dt_egreso = None

	try:
		dt_ingreso = datetime.combine( fecha_date, datetime.min.time()).replace(
			hour=hr_ing[0], minute=hr_ing[1]
		)
	except:
		dt_ingreso = None

	return (dt_reporte, dt_egreso, dt_ingreso)

#--------------------------------------------------------------------
#--------------------------------------------------------------------
def import_traslados_from_excel (file_path, sheet_name="ABRIL"):
	skipN  = 7		 # Number of lines to skip containing excel header
	from openpyxl import load_workbook
	wb = load_workbook (file_path, data_only=True)
	ws = wb[sheet_name]

	registros = []
	for row in ws.iter_rows(min_row=skipN + 2):  # +2 because Excel starts at 1
		try:
			values = []
			for cell in row:
				values.append(str(cell.value) if cell.value is not None else '')

			print (f"+++ {values=}")
			fechaReporte, fechaEgreso, fechaIngreso = parse_fechas_horas( values[0], values[1], values[2], values[3])
			print (f"\n+++ {fechaReporte=}, {fechaEgreso=}, {fechaIngreso=}")


			obj = Traslado (
				fecha_reporte=fechaReporte,
				fecha_egreso=fechaEgreso,
				fecha_ingreso=fechaIngreso,
				nombre_paciente=str (values[4]).strip (),
				documento=str (values[5]).strip (),
				servicio=str (values[6]).strip (),
				quien_reporta=str (values[7]).strip (),
				destino=str (values[8]).strip (),
				procedimiento=str (values[9]).strip (),
				medico=str (values[10]).strip (),
				aux_enfermeria=str (values[11]).strip (),
				conductor=str (values[12]).strip (),
				radio_operador=str (values[13]).strip (),
				ambulancia=str (values[14]).strip (),
				observacion="" if pd.isna (values[15]) else str (values[15]).strip (),
				mes=fechaReporte.month,
			)
		except Exception as ex:
			print (f"+++ Error leyendo archivo excel, fila: {row+skipN+2}, Mensaje: {str(ex)}")
			traceback.print_exc ()
			continue

		for field in obj._meta.fields:
			print (">>>", field.name, getattr(obj, field.name))
		registros.append (obj)

	# Bulk insert (fast)
	for r in registros:
		print (f"+++ {dir (r)=}")

	Traslado.objects.bulk_create (registros, batch_size=500)
	return len (registros)

#--------------------------------------------------------------------
#--------------------------------------------------------------------
import os, sys
import django

# --- DJANGO SETUP ---
# Add to Python path
PROJECT_PATH = "/home/lg/APPS/Hospital/hudn_crue_traslados/hudn_crue_traslados_app/"
sys.path.insert (0, PROJECT_PATH)
os.environ.setdefault ("DJANGO_SETTINGS_MODULE", "config.settings")
django.setup ()
from crue_traslados.models import Traslado

# --- IMPORT YOUR FUNCTION ---
#from traslados.utils import import_traslados_from_excel  # adjust path if needed

def main ():
	file_path = "traslados-abril.xlsx"	# change if needed
	sheet_name = "ABRIL"

	print ("Starting import...")

	count = import_traslados_from_excel (file_path, sheet_name)

	print (f"Import finished. {count} records inserted.")


if __name__ == "__main__":
	main ()
